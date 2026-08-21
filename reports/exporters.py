"""ساخت خروجی اکسل و PDF از گزارش‌ها.

هر گزارش به یک ساختار ساده تبدیل می‌شود (عنوان، سرستون‌ها، سطرها) و همین
ساختار هم به اکسل می‌رود هم به PDF. بنابراین اضافه کردن خروجی به یک گزارش
جدید فقط چند خط است.
"""
import io
from decimal import Decimal

from django.http import HttpResponse

from core.jalali import format_jalali, today_jalali_str
from core.money import format_amount

PERSIAN_DIGITS = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")


def _safe_filename(name):
    """نام فایل امن برای هدر HTTP (بدون کاراکتر غیر ASCII)."""
    keep = "-_. "
    ascii_name = "".join(c for c in name if c.isascii() and (c.isalnum() or c in keep)).strip()
    return ascii_name or "report"


# --------------------------------------------------------------------------
# اکسل
# --------------------------------------------------------------------------
def to_xlsx(title, headers, rows, *, subtitle="", column_widths=None, filename="report"):
    """یک فایل اکسل راست‌چین می‌سازد و به صورت پاسخ HTTP برمی‌گرداند."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "گزارش"
    sheet.sheet_view.rightToLeft = True

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    column_count = max(len(headers), 1)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.alignment = center

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
    meta = subtitle or f"تاریخ تهیه گزارش: {today_jalali_str().translate(PERSIAN_DIGITS)}"
    cell = sheet.cell(row=2, column=1, value=meta)
    cell.alignment = center
    cell.font = Font(size=9, color="666666")

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(rows, start=5):
        for col_index, value in enumerate(row, start=1):
            if isinstance(value, Decimal):
                value = float(value)
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = right
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.##"

    widths = column_widths or [18] * column_count
    for index, width in enumerate(widths[:column_count], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A5"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_safe_filename(filename)}.xlsx"; '
        f"filename*=UTF-8''{filename}.xlsx"
    )
    return response


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------
def _register_persian_font():
    """یک فونت فارسی برای PDF پیدا و ثبت می‌کند.

    فونت‌های ویندوز (Tahoma / Segoe UI) حروف فارسی را دارند. اگر هیچ‌کدام
    نبود، None برمی‌گردد و گزارش با فونت پیش‌فرض ساخته می‌شود.
    """
    import os

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        ("Vazirmatn", os.path.join(os.path.dirname(__file__), "fonts", "Vazirmatn-Regular.ttf")),
        ("Tahoma", r"C:\Windows\Fonts\tahoma.ttf"),
        ("Segoe UI", r"C:\Windows\Fonts\segoeui.ttf"),
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for name, path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue
    return None


def _shape_persian(text):
    """متن فارسی را برای PDF آماده می‌کند.

    ReportLab حروف فارسی را به هم نمی‌چسباند و ترتیب راست‌چین را هم رعایت
    نمی‌کند. اگر کتابخانه‌های arabic_reshaper و python-bidi نصب باشند از
    آن‌ها استفاده می‌کنیم؛ وگرنه متن معکوس می‌شود تا دست‌کم خوانا بماند.
    """
    if not text:
        return ""
    text = str(text)
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        return get_display(arabic_reshaper.reshape(text))
    except ImportError:
        has_persian = any("\u0600" <= ch <= "\u06ff" for ch in text)
        return text[::-1] if has_persian else text


def to_pdf(title, headers, rows, *, subtitle="", filename="report", landscape_mode=True):
    """گزارش را به صورت PDF برمی‌گرداند."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    font_name = _register_persian_font() or "Helvetica"
    page_size = landscape(A4) if landscape_mode else A4

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=page_size,
        rightMargin=12 * mm, leftMargin=12 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title=title,
    )

    title_style = ParagraphStyle("t", fontName=font_name, fontSize=14, alignment=1, spaceAfter=4)
    sub_style = ParagraphStyle("s", fontName=font_name, fontSize=8.5, alignment=1,
                               textColor=colors.HexColor("#666666"), spaceAfter=10)

    story = [
        Paragraph(_shape_persian(title), title_style),
        Paragraph(
            _shape_persian(subtitle or f"تاریخ تهیه: {today_jalali_str().translate(PERSIAN_DIGITS)}"),
            sub_style,
        ),
        Spacer(1, 4),
    ]

    # ستون‌ها برای راست‌چین بودن معکوس می‌شوند
    table_data = [[_shape_persian(h) for h in reversed(headers)]]
    for row in rows:
        table_data.append([_shape_persian(cell) for cell in reversed([str(c) for c in row])])

    table = Table(table_data, repeatRows=1, hAlign="CENTER")
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F766E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="{_safe_filename(filename)}.pdf"; '
        f"filename*=UTF-8''{filename}.pdf"
    )
    return response


# --------------------------------------------------------------------------
# کمکی‌های قالب‌بندی برای سطرهای گزارش
# --------------------------------------------------------------------------
def fmt_date(value):
    return format_jalali(value).translate(PERSIAN_DIGITS)


def fmt_number(value, decimal_places=0):
    if value is None:
        return "—"
    return format_amount(value, decimal_places).translate(PERSIAN_DIGITS)
