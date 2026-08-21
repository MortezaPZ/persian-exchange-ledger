"""انتقال داده از گوگل‌شیت / اکسل فعلی صرافی به سامانه.

فرمت پشتیبانی‌شده همان فرمت فایل کارفرماست: هر شیت یک مشتری، با ستون‌های
«تومن بدهکار/بستانکار»، «درهم بدهکار/بستانکار»، «دلار بدهکار/بستانکار»،
«تعداد ارز»، «قیمت فروش»، «قیمت خرید»، «شرح عملیات» و «تاریخ».

سه حالت اجرا:

    --mode check     فقط بررسی می‌کند و ایرادهای فایل را گزارش می‌دهد (چیزی ثبت نمی‌شود)
    --mode opening   مانده نهایی هر ارز را به صورت یک سند افتتاحیه ثبت می‌کند  ← پیشنهادی
    --mode full      تک‌تک سطرها را به عنوان سند وارد می‌کند

نمونه:
    python manage.py import_sheet "صورتحساب.xlsx" --mode check
    python manage.py import_sheet "صورتحساب.xlsx" --mode opening --user admin
"""
import re
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounts.models import User
from core.jalali import parse_jalali, today_gregorian
from core.money import parse_amount
from core.models import Currency, Party
from ledger import services

ZERO = Decimal("0")
EMPTY_MARKERS = {"", "-", "--", "---", "----", "-----", "------", "—", "؟"}

#: نگاشت عنوان ستون‌های فایل کارفرما به معنای آن‌ها
COLUMN_PATTERNS = [
    ("toman_credit", r"^(تومن|تومان|ریال)\s*بستانکار"),
    ("toman_debit", r"^(تومن|تومان|ریال)\s*بدهکار"),
    ("aed_credit", r"^درهم\s*بستانکار"),
    ("aed_debit", r"^درهم\s*بدهکار"),
    ("usd_credit", r"^دلار\s*بستانکار"),
    ("usd_debit", r"^دلار\s*بدهکار"),
    ("usdt_credit", r"^تتر\s*بستانکار"),
    ("usdt_debit", r"^تتر\s*بدهکار"),
    ("quantity", r"^تعداد\s*ارز"),
    ("sell_price", r"^(قیمت|نرخ)\s*فروش"),
    ("buy_price", r"^(قیمت|نرخ)\s*خرید"),
    ("description", r"^شرح"),
    ("date", r"^تاریخ"),
]

CURRENCY_BY_PREFIX = {"toman": "IRR", "aed": "AED", "usd": "USD", "usdt": "USDT"}


def normalize_header(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_number(value):
    """سلول را به عدد تبدیل می‌کند؛ «-----» و مشابهش یعنی خالی."""
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip()
    if text in EMPTY_MARKERS or not text:
        return None
    text = re.sub(r"[^\d۰-۹٠-٩.,+\-]", "", text)
    if not text or text in {"+", "-"}:
        return None
    try:
        return parse_amount(text)
    except ValueError:
        return None


class Command(BaseCommand):
    help = "انتقال داده از فایل اکسل/گوگل‌شیت فعلی به سامانه"

    def add_arguments(self, parser):
        parser.add_argument("path", help="مسیر فایل .xlsx")
        parser.add_argument("--mode", choices=["check", "opening", "full"], default="check")
        parser.add_argument("--user", default=None, help="نام کاربری ثبت‌کننده اسناد")
        parser.add_argument("--sheet", default=None, help="فقط همین شیت پردازش شود")
        parser.add_argument("--customer", default=None,
                            help="نام مشتری (پیش‌فرض: از نام شیت خوانده می‌شود)")

    def handle(self, *args, **options):
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"فایل پیدا نشد: {path}")

        try:
            import openpyxl
        except ImportError:
            raise CommandError("کتابخانه openpyxl نصب نیست: pip install openpyxl")

        mode = options["mode"]
        user = None
        if mode != "check":
            username = options["user"]
            if not username:
                raise CommandError("برای ثبت واقعی، نام کاربری ثبت‌کننده را با --user بدهید.")
            user = User.objects.filter(username=username).first()
            if user is None:
                raise CommandError(f"کاربری با نام «{username}» پیدا نشد.")

        workbook = openpyxl.load_workbook(path, data_only=True)
        sheets = [workbook[options["sheet"]]] if options["sheet"] else list(workbook)

        total_issues = 0
        for sheet in sheets:
            if sheet.max_row < 2:
                continue
            issues = self.process_sheet(sheet, mode, user, options.get("customer"))
            total_issues += issues

        self.stdout.write("")
        if total_issues:
            self.stdout.write(self.style.WARNING(f"مجموع {total_issues} مورد نیازمند بررسی."))
        else:
            self.stdout.write(self.style.SUCCESS("هیچ ایرادی پیدا نشد."))

    # ----------------------------------------------------------------------
    def process_sheet(self, sheet, mode, user, customer_name):
        self.stdout.write("")
        self.stdout.write(self.style.HTTP_INFO(f"═══ شیت: {sheet.title}"))

        header_row = [normalize_header(c.value) for c in sheet[1]]
        columns = {}
        for index, header in enumerate(header_row):
            for key, pattern in COLUMN_PATTERNS:
                if key not in columns and re.match(pattern, header):
                    columns[key] = index
                    break

        if "date" not in columns or "description" not in columns:
            self.stdout.write(self.style.WARNING("  ستون تاریخ یا شرح پیدا نشد؛ این شیت رد شد."))
            return 0

        currencies = {c.code: c for c in Currency.objects.all()}
        totals = {}
        issues = 0
        rows_parsed = 0

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() not in EMPTY_MARKERS for cell in row):
                continue

            def cell(key):
                idx = columns.get(key)
                return row[idx] if idx is not None and idx < len(row) else None

            description = str(cell("description") or "").strip()
            raw_date = cell("date")
            if not description and raw_date is None:
                continue

            rows_parsed += 1

            # --- جمع‌بندی اثر هر ارز در این سطر ---
            for prefix, code in CURRENCY_BY_PREFIX.items():
                currency = currencies.get(code)
                if currency is None:
                    continue
                debit = clean_number(cell(f"{prefix}_debit"))
                credit = clean_number(cell(f"{prefix}_credit"))
                effect = ZERO
                if debit:
                    effect += debit
                if credit:
                    effect -= credit
                if effect != ZERO:
                    totals[code] = totals.get(code, ZERO) + effect

            # --- بررسی سلامت: تعداد × نرخ باید با مبلغ ثبت‌شده بخواند ---
            quantity = clean_number(cell("quantity"))
            price = clean_number(cell("sell_price")) or clean_number(cell("buy_price"))
            recorded = clean_number(cell("toman_debit")) or clean_number(cell("toman_credit"))

            if quantity and price and recorded:
                expected = quantity * price
                if abs(expected - recorded) > Decimal("1"):
                    issues += 1
                    self.stdout.write(self.style.ERROR(
                        f"  ✗ سطر {row_index}: {quantity:g} × {price:g} = {expected:,.0f} "
                        f"ولی {recorded:,.0f} ثبت شده (اختلاف {expected - recorded:,.0f})"
                    ))

            if raw_date is not None:
                try:
                    parse_jalali(raw_date)
                except (ValueError, TypeError):
                    issues += 1
                    self.stdout.write(self.style.WARNING(
                        f"  ! سطر {row_index}: تاریخ «{raw_date}» خوانده نشد"
                    ))

        self.stdout.write(f"  {rows_parsed} سطر خوانده شد.")
        self.stdout.write("  مانده نهایی محاسبه‌شده از ستون‌های بدهکار/بستانکار:")
        for code, amount in totals.items():
            state = "بدهکار" if amount > 0 else "بستانکار"
            self.stdout.write(f"    · {currencies[code].name}: {abs(amount):,.2f} {state}")

        if mode == "opening":
            issues += self.post_opening(sheet, totals, currencies, user, customer_name)
        elif mode == "full":
            self.stdout.write(self.style.WARNING(
                "  حالت full هنوز فعال نیست چون سطرهای فایل فعلی نوع تراکنش مشخصی ندارند.\n"
                "  از حالت opening استفاده کنید و معاملات جدید را در خود سامانه ثبت کنید."
            ))

        return issues

    # ----------------------------------------------------------------------
    @transaction.atomic
    def post_opening(self, sheet, totals, currencies, user, customer_name):
        """مانده نهایی هر ارز را به صورت سند افتتاحیه ثبت می‌کند."""
        name = customer_name or re.sub(r"^صورتحساب\s*", "", sheet.title).strip() or sheet.title

        party, created = Party.objects.get_or_create(
            kind=Party.Kind.CUSTOMER, name=name,
            defaults={"note": f"از فایل «{sheet.title}» منتقل شد."},
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"  ✓ مشتری «{party.name}» ساخته شد"))
        else:
            self.stdout.write(f"  · مشتری «{party.name}» از قبل وجود داشت")

        issues = 0
        today = today_gregorian()
        for code, amount in totals.items():
            if amount == ZERO:
                continue
            currency = currencies[code]
            external_key = f"opening:{party.pk}:{code}"
            try:
                voucher = services.post_opening_balance(
                    date=today, party=party, currency=currency, amount=amount,
                    created_by=user,
                    description=f"مانده افتتاحیه منتقل‌شده از فایل «{sheet.title}»",
                    external_key=external_key,
                )
            except services.DuplicateVoucher as exc:
                self.stdout.write(f"    · {currency.name}: قبلاً ثبت شده (سند {exc.voucher.number})")
            except services.LedgerError as exc:
                issues += 1
                self.stdout.write(self.style.ERROR(f"    ✗ {currency.name}: {exc}"))
            else:
                self.stdout.write(self.style.SUCCESS(
                    f"    ✓ {currency.name}: سند افتتاحیه {voucher.number} ثبت شد"
                ))
        return issues
